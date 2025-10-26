from typing import Dict, Any, Tuple, List, Optional, Generator
import html
import time
import string
import openpyxl
from io import BytesIO, StringIO
from fastapi.responses import JSONResponse
from copy import deepcopy
import multiprocessing as mp

from spiderfoot import SpiderFootDb
from configs import SF_CONFIG, logger


# --- Utility Function to Process Data Rows (Common to CSV and Excel) ---
def process_data(data: List[tuple]) -> Generator[List[Any], None, None]:
    """Processes the raw database rows into the desired output format."""
    for row in data:
        # Skip 'ROOT' type rows
        if row[4] == "ROOT":
            continue

        # Format timestamp (row[0] is assumed to be a timestamp in seconds)
        lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))

        # Clean up data field (row[1])
        datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")

        # Return the structured row:
        yield [lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield]

# --- Common Data Processing ---
def process_multi_data(data: List[tuple], scaninfo: Dict[str, Tuple[str, ...]]) -> Generator[List[Any], None, None]:
    """Processes raw DB rows into the structured list format used for export, including scan name."""
    for row in data:
        # Skip 'ROOT' type rows
        if row[4] == "ROOT":
            continue

        # Get the scan ID and name
        scan_id = str(row[12])
        scan_name = scaninfo.get(scan_id, (None,))[0] or "N/A"

        # Format timestamp (row[0] is assumed to be a timestamp in seconds)
        lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))

        # Clean up data field (row[1])
        datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")

        # Yield the structured row:
        # [Scan Name, Updated, Type, Module, Source, F/P, Data]
        # (row[12] is the scan ID, row[4] is Type, row[3] is Module, row[2] is Source, row[13] is F/P)
        yield [scan_name, lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield]

def buildExcel(data: list, columnNames: list, sheetNameIndex: int = 0) -> bytes:
    """raw 데이터를 Excel 워크북으로 변환합니다 (기존 로직 유지)."""
    rowNums = dict()
    workbook = openpyxl.Workbook()
    defaultSheet = workbook.active
    
    # pop을 위해 복사본 사용
    col_titles = columnNames[:]
    sheet_name_col = col_titles.pop(sheetNameIndex)
    
    allowed_sheet_chars = string.ascii_uppercase + string.digits + '_'
    
    for row_data in data:
        # pop을 위해 복사본 사용
        row = row_data[:] 
        
        # sheetNameIndex를 기반으로 시트 이름을 가져옴
        sheet_name_value = str(row.pop(sheetNameIndex))
        sheetName = "".join([c for c in sheet_name_value if c.upper() in allowed_sheet_chars])
        
        try:
            sheet = workbook[sheetName]
        except KeyError:
            workbook.create_sheet(sheetName)
            sheet = workbook[sheetName]
            
            for col_num, column_title in enumerate(col_titles, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.value = column_title
            rowNums[sheetName] = 2

        for col_num, cell_value in enumerate(row, 1):
            cell = sheet.cell(row=rowNums[sheetName], column=col_num)
            cell.value = cell_value

        rowNums[sheetName] += 1

    if rowNums:
        workbook.remove(defaultSheet)

    workbook._sheets.sort(key=lambda ws: ws.title)

    with BytesIO() as f:
        workbook.save(f)
        f.seek(0)
        return f.read()

def reset_settings(dbh) -> bool:
    """Reset settings to default.

    Returns:
        bool: success
    """
    try:
        dbh.configClear()  
        state.config = deepcopy(SF_CONFIG) 
        return True 
    except Exception as e:
        logger.error(f'reset settings: {e}')
        return False

def jsonify_error(status: int, message: str) -> JSONResponse:
    """JSON 형식의 에러 응답을 반환합니다."""
    return JSONResponse(
        status_code=status,
        content={'message': message}
    )